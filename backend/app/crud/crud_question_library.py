"""
试题库功能相关的CRUD操作
"""

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, and_, func
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
import json
import tempfile
import os
import uuid

from app.models import models
from app.schemas import schemas


def get_question_library_list(
    db: Session,
    teacher_id: int,
    keyword: Optional[str] = None,
    question_type: Optional[str] = None,
    direction: Optional[str] = None,
    category: Optional[str] = None,
    difficulty: Optional[str] = None,
    source: Optional[str] = None,
    skip: int = 0,
    limit: int = 20
):
    """获取试题库列表"""
    # 基础查询
    query = db.query(models.Question).filter(
        or_(
            models.Question.creator_id == teacher_id,  # 个人试题
            models.Question.is_shared == True  # 共享试题
        )
    )
    
    # 关键词搜索（搜索题干内容）
    if keyword:
        query = query.filter(models.Question.content.ilike(f"%{keyword}%"))
    
    # 题目类型筛选
    if question_type:
        query = query.filter(models.Question.question_type == question_type)
    
    # 方向筛选
    if direction:
        query = query.filter(models.Question.direction == direction)
    
    # 分类筛选
    if category:
        query = query.filter(models.Question.category == category)
    
    # 难度筛选
    if difficulty:
        query = query.filter(models.Question.difficulty == difficulty)
    
    # 来源筛选
    if source:
        if source == "个人":
            query = query.filter(models.Question.creator_id == teacher_id)
        elif source == "平台":
            query = query.filter(models.Question.is_shared == True)
    
    # 获取总数
    total = query.count()
    
    # 分页查询
    questions = query.order_by(
        models.Question.created_at.desc()
    ).offset(skip).limit(limit).all()
    
    # 构建响应数据
    question_list = []
    for question in questions:
        # 判断来源
        source_type = "个人" if question.creator_id == teacher_id else "平台"
        
        # 题型中文显示
        question_type_cn_map = {
            "SINGLE_CHOICE": "单选题",
            "MULTIPLE_CHOICE": "多选题",
            "TRUE_FALSE": "判断题",
            "SHORT_ANSWER": "简答题"
        }
        question_type_cn = question_type_cn_map.get(question.question_type.value if question.question_type else None, "未知")
        
        # 难度中文显示
        difficulty_cn_map = {
            "BEGINNER": "初级",
            "INTERMEDIATE": "中级", 
            "ADVANCED": "高级"
        }
        difficulty_cn = difficulty_cn_map.get(question.difficulty.value if question.difficulty else None, "未设置")
        
        # 题干摘要（截取前100字符）
        content_summary = question.content[:100] + "..." if len(question.content) > 100 else question.content
        
        # 获取创建者信息
        creator_name = "系统"
        if question.creator_id:
            creator = db.query(models.User).filter(models.User.id == question.creator_id).first()
            creator_name = creator.full_name if creator else "系统"

        question_data = {
            "id": question.id,
            "content": content_summary,
            "question_type": question.question_type.value if question.question_type else None,
            "question_type_cn": question_type_cn,
            "difficulty": question.difficulty.value if question.difficulty else None,
            "difficulty_cn": difficulty_cn,
            "direction": question.direction,
            "category": question.category,
            "source": source_type,
            "creator_name": creator_name,
            "created_at": question.created_at,
            "updated_at": question.updated_at,
            "can_copy": True,
            "can_edit": question.creator_id == teacher_id,
            "can_delete": question.creator_id == teacher_id
        }
        question_list.append(question_data)
    
    return question_list, total


def create_question(
    db: Session,
    question_data: dict,
    creator_id: int
):
    """创建试题"""
    # 处理选项数据
    options_json = None
    if question_data.get("options"):
        # 如果选项已经是字典格式，直接使用；否则调用model_dump()
        if isinstance(question_data["options"][0], dict):
            options_json = json.dumps(question_data["options"], ensure_ascii=False)
        else:
            options_json = json.dumps([opt.model_dump() for opt in question_data["options"]], ensure_ascii=False)
    
    # 处理正确答案
    correct_answers_json = None
    if question_data.get("correct_answers"):
        correct_answers_json = json.dumps(question_data["correct_answers"], ensure_ascii=False)
    
    # 生成唯一ID
    question_id = f"q_{uuid.uuid4().hex[:12]}"

    # 创建试题记录
    new_question = models.Question(
        id=question_id,
        content=question_data["content"],
        question_type=question_data["question_type"],
        options=options_json,
        correct_answers=correct_answers_json,
        explanation=question_data.get("explanation"),
        difficulty=question_data.get("difficulty"),
        direction=question_data.get("direction"),
        category=question_data.get("category"),
        creator_id=creator_id,
        is_shared=False  # 个人创建的试题默认不共享
    )
    
    db.add(new_question)
    db.commit()
    db.refresh(new_question)
    
    return new_question


def copy_question(
    db: Session,
    question_id: str,
    teacher_id: int
):
    """复制试题"""
    # 获取原试题
    original_question = db.query(models.Question).filter(
        models.Question.id == question_id
    ).first()
    
    if not original_question:
        return None
    
    # 创建新试题
    new_question = models.Question(
        content=original_question.content,
        question_type=original_question.question_type,
        options=original_question.options,
        correct_answers=original_question.correct_answers,
        explanation=original_question.explanation,
        difficulty=original_question.difficulty,
        direction=original_question.direction,
        category=original_question.category,
        creator_id=teacher_id,
        is_shared=False
    )
    
    db.add(new_question)
    db.commit()
    db.refresh(new_question)
    
    return new_question


def update_question(
    db: Session,
    question_id: str,
    teacher_id: int,
    update_data: dict
):
    """更新试题（个人试题直接编辑，内置试题自动复制为个人试题）"""
    # 获取试题
    question = db.query(models.Question).filter(
        models.Question.id == question_id
    ).first()
    
    if not question:
        return None
    
    # 如果是个人试题，直接编辑
    if question.creator_id == teacher_id:
        # 更新字段
        if "content" in update_data and update_data["content"] is not None:
            question.content = update_data["content"]
        
        if "options" in update_data and update_data["options"] is not None:
            # 处理选项数据
            if isinstance(update_data["options"][0], dict):
                question.options = json.dumps(update_data["options"], ensure_ascii=False)
            else:
                question.options = json.dumps([opt.model_dump() for opt in update_data["options"]], ensure_ascii=False)
        
        if "correct_answers" in update_data and update_data["correct_answers"] is not None:
            question.correct_answers = json.dumps(update_data["correct_answers"], ensure_ascii=False)
        
        if "explanation" in update_data:
            question.explanation = update_data["explanation"]
        
        if "difficulty" in update_data and update_data["difficulty"] is not None:
            question.difficulty = update_data["difficulty"]
        
        if "direction" in update_data:
            question.direction = update_data["direction"]
        
        if "category" in update_data:
            question.category = update_data["category"]
        
        question.updated_at = datetime.now(timezone.utc)
        
        db.commit()
        db.refresh(question)
        
        return question
    
    # 如果是内置试题（共享试题），先复制为个人试题，再编辑
    elif question.is_shared:
        # 处理选项数据
        options_json = None
        if update_data.get("options"):
            if isinstance(update_data["options"][0], dict):
                options_json = json.dumps(update_data["options"], ensure_ascii=False)
            else:
                options_json = json.dumps([opt.model_dump() for opt in update_data["options"]], ensure_ascii=False)
        else:
            options_json = question.options
        
        # 处理正确答案
        correct_answers_json = None
        if update_data.get("correct_answers"):
            correct_answers_json = json.dumps(update_data["correct_answers"], ensure_ascii=False)
        else:
            correct_answers_json = question.correct_answers
        
        # 创建新的个人试题
        new_question = models.Question(
            content=update_data.get("content", question.content),
            question_type=question.question_type,
            options=options_json,
            correct_answers=correct_answers_json,
            explanation=update_data.get("explanation", question.explanation),
            difficulty=update_data.get("difficulty", question.difficulty),
            direction=update_data.get("direction", question.direction),
            category=update_data.get("category", question.category),
            creator_id=teacher_id,
            is_shared=False
        )
        
        db.add(new_question)
        db.commit()
        db.refresh(new_question)
        
        return new_question
    
    # 无权限编辑
    else:
        return None


def delete_question(
    db: Session,
    question_id: str,
    teacher_id: int
):
    """删除试题（仅个人试题）"""
    # 获取试题
    question = db.query(models.Question).filter(
        models.Question.id == question_id,
        models.Question.creator_id == teacher_id  # 只能删除自己创建的试题
    ).first()
    
    if not question:
        return False
    
    # 检查是否被试卷使用
    paper_question_count = db.query(models.TestPaperQuestion).filter(
        models.TestPaperQuestion.question_id == question_id
    ).count()
    
    if paper_question_count > 0:
        # 如果被试卷使用，不能删除
        return False
    
    # 删除试题
    db.delete(question)
    db.commit()
    
    return True


def batch_delete_questions(
    db: Session,
    question_ids: List[str],
    teacher_id: int
):
    """批量删除试题"""
    deleted_count = 0
    failed_questions = []
    
    for question_id in question_ids:
        success = delete_question(db, question_id, teacher_id)
        if success:
            deleted_count += 1
        else:
            failed_questions.append(question_id)
    
    return deleted_count, failed_questions


def get_question_detail(
    db: Session,
    question_id: str,
    teacher_id: int
):
    """获取试题详情"""
    question = db.query(models.Question).filter(
        models.Question.id == question_id,
        or_(
            models.Question.creator_id == teacher_id,  # 个人试题
            models.Question.is_shared == True  # 共享试题
        )
    ).first()
    
    if not question:
        return None
    
    # 解析选项数据
    options = []
    if question.options:
        try:
            options_data = json.loads(question.options)
            options = [{"key": opt["key"], "content": opt["content"]} for opt in options_data]
        except (json.JSONDecodeError, KeyError):
            options = []

    # 解析正确答案
    correct_answers = []
    if question.correct_answers:
        try:
            correct_answers = json.loads(question.correct_answers)
        except json.JSONDecodeError:
            correct_answers = []
    
    # 判断来源
    source_type = "个人" if question.creator_id == teacher_id else "平台"
    
    # 题型中文显示
    question_type_cn_map = {
        "SINGLE_CHOICE": "单选题",
        "MULTIPLE_CHOICE": "多选题",
        "TRUE_FALSE": "判断题",
        "SHORT_ANSWER": "简答题"
    }
    question_type_cn = question_type_cn_map.get(question.question_type.value if question.question_type else None, "未知")
    
    # 难度中文显示
    difficulty_cn_map = {
        "BEGINNER": "初级",
        "INTERMEDIATE": "中级", 
        "ADVANCED": "高级"
    }
    difficulty_cn = difficulty_cn_map.get(question.difficulty.value if question.difficulty else None, "未设置")
    
    # 获取创建者信息（手动查询，避免relationship加载问题）
    creator_name = "系统"
    if question.creator_id:
        creator = db.query(models.User).filter(models.User.id == question.creator_id).first()
        creator_name = creator.full_name if creator else "系统"
    
    question_detail = {
        "id": question.id,
        "content": question.content,
        "question_type": question.question_type.value if question.question_type else None,
        "question_type_cn": question_type_cn,
        "options": options,
        "correct_answers": correct_answers,
        "explanation": question.explanation,
        "difficulty": question.difficulty.value if question.difficulty else None,
        "difficulty_cn": difficulty_cn,
        "direction": question.direction,
        "category": question.category,
        "source": source_type,
        "creator_id": question.creator_id,
        "creator_name": creator_name,
        "created_at": question.created_at,
        "updated_at": question.updated_at,
        "can_edit": question.creator_id == teacher_id,
        "can_delete": question.creator_id == teacher_id
    }
    
    return question_detail


def get_question_filter_tags(db: Session):
    """获取试题筛选标签"""
    # 题型标签
    question_types = [
        {"value": "SINGLE_CHOICE", "label": "单选题"},
        {"value": "MULTIPLE_CHOICE", "label": "多选题"},
        {"value": "TRUE_FALSE", "label": "判断题"},
        {"value": "SHORT_ANSWER", "label": "简答题"}
    ]
    
    # 方向标签
    directions = db.query(models.Question.direction).filter(
        models.Question.direction.isnot(None),
        models.Question.direction != ""
    ).distinct().all()
    directions = [d[0] for d in directions if d[0]]
    
    # 分类标签
    categories = db.query(models.Question.category).filter(
        models.Question.category.isnot(None),
        models.Question.category != ""
    ).distinct().all()
    categories = [c[0] for c in categories if c[0]]
    
    # 难度标签
    difficulties = [
        {"value": "BEGINNER", "label": "初级"},
        {"value": "INTERMEDIATE", "label": "中级"},
        {"value": "ADVANCED", "label": "高级"}
    ]
    
    # 来源标签
    sources = ["个人", "平台"]
    
    return {
        "question_types": question_types,
        "directions": directions,
        "categories": categories,
        "difficulties": difficulties,
        "sources": sources
    }


def import_questions_from_file(
    db: Session,
    file_path: str,
    teacher_id: int
):
    """从文件导入试题"""
    success_count = 0
    error_count = 0
    error_details = []
    
    try:
        # 判断文件类型
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext in ['.xlsx', '.xls']:
            # Excel文件
            try:
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path, read_only=True)
                
                # 难度映射
                difficulty_map = {
                    '初级': 'BEGINNER',
                    '中级': 'INTERMEDIATE',
                    '高级': 'ADVANCED'
                }
                
                # 处理单选题
                if '单选题' in wb.sheetnames:
                    ws = wb['单选题']
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                        if not row[0]:  # 题干为空，跳过
                            continue
                        
                        try:
                            options = []
                            for i in range(1, 7):  # 选项A-F
                                if row[i]:
                                    options.append({
                                        "key": chr(64 + i),  # A, B, C...
                                        "content": str(row[i])
                                    })
                            
                            correct_answer = str(row[7]).upper() if row[7] else ""
                            
                            question_data = {
                                "content": str(row[0]),
                                "question_type": "SINGLE_CHOICE",
                                "options": options,
                                "correct_answers": [correct_answer],
                                "explanation": str(row[8]) if row[8] else None,
                                "direction": str(row[9]) if row[9] else None,
                                "category": str(row[10]) if row[10] else None,
                                "difficulty": difficulty_map.get(str(row[11]), "INTERMEDIATE") if row[11] else "INTERMEDIATE"
                            }
                            
                            create_question(db, question_data, teacher_id)
                            success_count += 1
                            
                        except Exception as e:
                            error_details.append({
                                "sheet": "单选题",
                                "row": row_idx,
                                "error": str(e)
                            })
                            error_count += 1
                
                # 处理多选题
                if '多选题' in wb.sheetnames:
                    ws = wb['多选题']
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                        if not row[0]:
                            continue
                        
                        try:
                            options = []
                            for i in range(1, 7):
                                if row[i]:
                                    options.append({
                                        "key": chr(64 + i),
                                        "content": str(row[i])
                                    })
                            
                            # 解析多选答案（如"ABD"）
                            correct_answers = []
                            if row[7]:
                                for char in str(row[7]).upper():
                                    if char in 'ABCDEF':
                                        correct_answers.append(char)
                            
                            question_data = {
                                "content": str(row[0]),
                                "question_type": "MULTIPLE_CHOICE",
                                "options": options,
                                "correct_answers": correct_answers,
                                "explanation": str(row[8]) if row[8] else None,
                                "direction": str(row[9]) if row[9] else None,
                                "category": str(row[10]) if row[10] else None,
                                "difficulty": difficulty_map.get(str(row[11]), "INTERMEDIATE") if row[11] else "INTERMEDIATE"
                            }
                            
                            create_question(db, question_data, teacher_id)
                            success_count += 1
                            
                        except Exception as e:
                            error_details.append({
                                "sheet": "多选题",
                                "row": row_idx,
                                "error": str(e)
                            })
                            error_count += 1
                
                # 处理判断题
                if '判断题' in wb.sheetnames:
                    ws = wb['判断题']
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                        if not row[0]:
                            continue
                        
                        try:
                            # 解析答案
                            answer = "true" if str(row[1]) in ["正确", "对", "是", "TRUE", "true", "True"] else "false"
                            
                            question_data = {
                                "content": str(row[0]),
                                "question_type": "TRUE_FALSE",
                                "correct_answers": [answer],
                                "explanation": str(row[2]) if row[2] else None,
                                "direction": str(row[3]) if row[3] else None,
                                "category": str(row[4]) if row[4] else None,
                                "difficulty": difficulty_map.get(str(row[5]), "INTERMEDIATE") if row[5] else "INTERMEDIATE"
                            }
                            
                            create_question(db, question_data, teacher_id)
                            success_count += 1
                            
                        except Exception as e:
                            error_details.append({
                                "sheet": "判断题",
                                "row": row_idx,
                                "error": str(e)
                            })
                            error_count += 1
                
                # 处理简答题
                if '简答题' in wb.sheetnames:
                    ws = wb['简答题']
                    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                        if not row[0]:
                            continue
                        
                        try:
                            question_data = {
                                "content": str(row[0]),
                                "question_type": "SHORT_ANSWER",
                                "correct_answers": [str(row[1])] if row[1] else ["暂无"],
                                "explanation": str(row[2]) if row[2] else None,
                                "direction": str(row[3]) if row[3] else None,
                                "category": str(row[4]) if row[4] else None,
                                "difficulty": difficulty_map.get(str(row[5]), "INTERMEDIATE") if row[5] else "INTERMEDIATE"
                            }
                            
                            create_question(db, question_data, teacher_id)
                            success_count += 1
                            
                        except Exception as e:
                            error_details.append({
                                "sheet": "简答题",
                                "row": row_idx,
                                "error": str(e)
                            })
                            error_count += 1
                
                wb.close()
                
            except ImportError:
                error_details.append({
                    "row": 0,
                    "error": "服务器缺少openpyxl库，无法处理Excel文件"
                })
                error_count = 1
                
        elif file_ext == '.json':
            # JSON文件
            with open(file_path, 'r', encoding='utf-8') as f:
                questions_data = json.load(f)
            
            for i, question_data in enumerate(questions_data):
                try:
                    # 验证必要字段
                    if not question_data.get("content"):
                        error_details.append({
                            "row": i + 1,
                            "error": "题干内容不能为空"
                        })
                        error_count += 1
                        continue
                    
                    if not question_data.get("question_type"):
                        error_details.append({
                            "row": i + 1,
                            "error": "题目类型不能为空"
                        })
                        error_count += 1
                        continue
                    
                    # 创建试题
                    create_question(db, question_data, teacher_id)
                    success_count += 1
                    
                except Exception as e:
                    error_details.append({
                        "row": i + 1,
                        "error": str(e)
                    })
                    error_count += 1
        else:
            error_details.append({
                "row": 0,
                "error": f"不支持的文件格式: {file_ext}"
            })
            error_count = 1
    
    except Exception as e:
        error_details.append({
            "row": 0,
            "error": f"文件解析失败: {str(e)}"
        })
        error_count = 1
    
    total_count = success_count + error_count
    
    return {
        "success_count": success_count,
        "error_count": error_count,
        "total_count": total_count,
        "error_details": error_details,
        "success": error_count == 0,
        "message": f"导入完成，成功{success_count}道，失败{error_count}道"
    }


def download_question_template():
    """下载试题导入模板"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # 创建工作簿
        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet
        
        # 定义表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 创建单选题sheet
        ws_single = wb.create_sheet("单选题")
        single_headers = ["题干*", "选项A*", "选项B*", "选项C", "选项D", "选项E", "选项F", 
                         "正确答案*", "解析", "方向分类*", "二级分类", "难度*"]
        for col, header in enumerate(single_headers, 1):
            cell = ws_single.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加示例数据
        ws_single.append([
            "Python是一种什么类型的编程语言？",
            "编译型", "解释型", "汇编型", "机器型", "", "",
            "B",
            "Python是一种解释型编程语言，代码在运行时由解释器逐行执行。",
            "Python基础", "编程语言", "初级"
        ])
        
        # 创建多选题sheet
        ws_multiple = wb.create_sheet("多选题")
        multiple_headers = ["题干*", "选项A*", "选项B*", "选项C", "选项D", "选项E", "选项F", 
                           "正确答案*", "解析", "方向分类*", "二级分类", "难度*"]
        for col, header in enumerate(multiple_headers, 1):
            cell = ws_multiple.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加示例数据
        ws_multiple.append([
            "以下哪些是Python的内置数据类型？",
            "list", "dict", "array", "tuple", "", "",
            "ABD",
            "list、dict、tuple是Python的内置数据类型，array需要导入numpy模块。",
            "Python基础", "数据类型", "中级"
        ])
        
        # 创建判断题sheet
        ws_judge = wb.create_sheet("判断题")
        judge_headers = ["题干*", "正确答案*", "解析", "方向分类*", "二级分类", "难度*"]
        for col, header in enumerate(judge_headers, 1):
            cell = ws_judge.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加示例数据
        ws_judge.append([
            "Python中的变量需要声明类型。",
            "错误",
            "Python是动态类型语言，变量不需要声明类型，类型在运行时确定。",
            "Python基础", "变量类型", "初级"
        ])
        
        # 创建简答题sheet
        ws_essay = wb.create_sheet("简答题")
        essay_headers = ["题干*", "参考答案*", "解析", "方向分类*", "二级分类", "难度*"]
        for col, header in enumerate(essay_headers, 1):
            cell = ws_essay.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 添加示例数据
        ws_essay.append([
            "请简述Python的特点。",
            "Python是一种解释型、面向对象、动态数据类型的高级程序设计语言。具有简单易学、可读性强、跨平台、丰富的库支持等特点。",
            "可以从语法简洁、动态类型、解释执行、跨平台、丰富的标准库等方面展开说明。",
            "Python基础", "语言特性", "中级"
        ])
        
        # 调整列宽
        for ws in [ws_single, ws_multiple, ws_judge, ws_essay]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到临时文件
        temp_dir = tempfile.gettempdir()
        filename = f"试题导入模板_{datetime.now().strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        wb.save(filepath)
        
        return {
            "filename": filename,
            "filepath": filepath,
            "success": True
        }
        
    except ImportError:
        # 如果没有openpyxl，返回JSON模板
        template_data = [
            {
                "content": "Python是一种什么类型的编程语言？",
                "question_type": "SINGLE_CHOICE",
                "options": [
                    {"key": "A", "content": "编译型"},
                    {"key": "B", "content": "解释型"},
                    {"key": "C", "content": "汇编型"},
                    {"key": "D", "content": "机器型"}
                ],
                "correct_answers": ["B"],
                "explanation": "Python是一种解释型编程语言，代码在运行时由解释器逐行执行。",
                "difficulty": "BEGINNER",
                "direction": "Python基础",
                "category": "编程语言"
            },
            {
                "content": "以下哪些是Python的内置数据类型？",
                "question_type": "MULTIPLE_CHOICE",
                "options": [
                    {"key": "A", "content": "list"},
                    {"key": "B", "content": "dict"},
                    {"key": "C", "content": "array"},
                    {"key": "D", "content": "tuple"}
                ],
                "correct_answers": ["A", "B", "D"],
                "explanation": "list、dict、tuple是Python的内置数据类型，array需要导入numpy模块。",
                "difficulty": "INTERMEDIATE",
                "direction": "Python基础",
                "category": "数据类型"
            },
            {
                "content": "Python中的变量需要声明类型。",
                "question_type": "TRUE_FALSE",
                "correct_answers": ["false"],
                "explanation": "Python是动态类型语言，变量不需要声明类型，类型在运行时确定。",
                "difficulty": "BEGINNER",
                "direction": "Python基础",
                "category": "变量类型"
            }
        ]
    
    # 创建临时文件
    temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8')
    json.dump(template_data, temp_file, ensure_ascii=False, indent=2)
    temp_file.close()
    
    return temp_file.name 